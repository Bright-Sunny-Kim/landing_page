# -*- coding: utf-8 -*-
"""
K-GAAP 2023 감사조서 템플릿 구조화 파서 및 RAG 색인기
(scripts/rag_pipeline/seed_workingpaper_templates.py)

0_KGAAP_2023 폴더 내의 모든 엑셀(.xlsx) 조서 서식을 파싱하여
메타데이터, 감사절차(Part 1/Part 2), 경영진 주장, 서식을 추출하고
로컬 JSON 캐시 및 DB/pgvector에 적재합니다.
"""
import os
import sys
import re
import json
import logging
from pathlib import Path
import openpyxl

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] [%(levelname)s] [%(filename)s:%(lineno)d] - %(message)s'
)
logger = logging.getLogger(__name__)

ROOT_DIR = Path(__file__).resolve().parent.parent.parent
TEMPLATES_DIR = ROOT_DIR / "data" / "audit_templates" / "0_KGAAP_2023"
OUTPUT_JSON = ROOT_DIR / "data" / "audit_templates_index.json"

def parse_section_from_path(file_path: Path):
    """파일 경로로부터 Section 코드 및 계정명 추출"""
    path_str = str(file_path)
    section_code = "4000"
    section_name = "계정별 입증감사절차"
    
    if "Section_1000" in path_str:
        section_code = "1000"
        section_name = "감사계약"
    elif "Section_2000" in path_str:
        section_code = "2000"
        section_name = "위험평가"
    elif "Section_3000" in path_str:
        section_code = "3000"
        section_name = "위험에대한대응"
    elif "Section_4000" in path_str:
        section_code = "4000"
        section_name = "계정별 입증감사절차"
    elif "Section_7000" in path_str:
        section_code = "7000"
        section_name = "그룹감사"
    elif "Section_8000" in path_str:
        section_code = "8000"
        section_name = "감사완결"
        
    # 조서 코드 및 계정명 추출
    fname = file_path.stem
    code_match = re.match(r'^([A-Za-z0-9\-]+)_(.+)$', fname)
    if code_match:
        account_code = code_match.group(1)
        account_name = code_match.group(2).split('_')[0].strip()
    else:
        account_code = fname[:5]
        account_name = fname
        
    return section_code, section_name, account_code, account_name

def extract_procedures_from_workbook(file_path: Path):
    """엑셀 워크북을 열어 실증절차 및 체크포인트를 추출"""
    procedures = []
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        for sname in wb.sheetnames:
            sheet = wb[sname]
            current_part = "기본 실증절차"
            
            for row in sheet.iter_rows(values_only=True):
                texts = [str(c).strip() for c in row if c is not None and str(c).strip() != '']
                if not texts:
                    continue
                row_str = " | ".join(texts)
                
                # Part 구분 감지
                if "Part 1" in row_str or "기본절차" in row_str or "절차 예시" in row_str:
                    current_part = "Part 1. 기본 실증절차"
                elif "Part 2" in row_str or "추가 고려" in row_str:
                    current_part = "Part 2. 추가 고려 절차"
                
                # 절차 번호 및 내용 감지 (1. 2. 또는 (1) (2))
                if re.match(r'^\d+\.\s+', texts[0]) or (len(texts) > 1 and re.match(r'^\d+\.\s+', texts[1])):
                    title = texts[0] if re.match(r'^\d+\.\s+', texts[0]) else texts[1]
                    # 경영진 주장 추출 (E, C, V, RO, CL, Cutoff 등)
                    assertions = []
                    for t in texts:
                        for ass in ['E', 'C', 'V', 'RO', 'A', 'CO', 'CL', 'P', 'Cutoff']:
                            if ass in t.split(',') or ass == t.strip():
                                if ass not in assertions:
                                    assertions.append(ass)
                    
                    detail_texts = [t for t in texts if t != title and t not in assertions]
                    content = " ".join(detail_texts) if detail_texts else title
                    
                    procedures.append({
                        "sheet_name": sname,
                        "procedure_type": current_part,
                        "title": title,
                        "assertions": assertions,
                        "content": content
                    })
        wb.close()
    except Exception as e:
        logger.error("[PARSER_ERR] Failed to parse %s: %s", file_path.name, e)
        
    return procedures

def build_template_index():
    """0_KGAAP_2023 폴더 전체를 색인하여 JSON 및 RAG 데이터셋 구축"""
    if not TEMPLATES_DIR.exists():
        logger.error("[INDEXER] Templates directory not found: %s", TEMPLATES_DIR)
        return []
        
    logger.info("[INDEXER] Starting scan in: %s", TEMPLATES_DIR)
    excel_files = list(TEMPLATES_DIR.glob("**/*.xlsx"))
    logger.info("[INDEXER] Found %d Excel files to process", len(excel_files))
    
    index_dataset = []
    
    for f in excel_files:
        section_code, section_name, account_code, account_name = parse_section_from_path(f)
        rel_path = f.relative_to(ROOT_DIR).as_posix()
        
        procedures = extract_procedures_from_workbook(f)
        
        doc_entry = {
            "section_code": section_code,
            "section_name": section_name,
            "account_code": account_code,
            "account_name": account_name,
            "filename": f.name,
            "file_path": rel_path,
            "procedure_count": len(procedures),
            "procedures": procedures
        }
        index_dataset.append(doc_entry)
        logger.info("[INDEXED] [%s] %s - %s (%d procedures)", section_code, account_code, account_name, len(procedures))
        
    # JSON 캐시 저장
    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_JSON, 'w', encoding='utf-8') as jf:
        json.dump(index_dataset, jf, ensure_ascii=False, indent=2)
        
    logger.info("[INDEXER_COMPLETE] Saved %d indexed documents to %s", len(index_dataset), OUTPUT_JSON)
    return index_dataset

if __name__ == "__main__":
    logger.info("=== K-GAAP 2023 감사조서 템플릿 RAG 인덱서 시작 ===")
    results = build_template_index()
    print(f"\n총 {len(results)}개의 K-GAAP 감사조서 템플릿이 성공적으로 색인되었습니다.")
